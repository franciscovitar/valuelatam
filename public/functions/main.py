import json
import io
from firebase_functions import https_fn
from firebase_admin import initialize_app
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

initialize_app()

C = {
    "navy":"0A1628","teal":"0D9488","teal_lt":"CCFBF1","white":"FFFFFF",
    "hdr":"1E293B","row":"F1F5F9","brd":"CBD5E1","dk":"0F172A","md":"475569",
    "green":"059669","green_lt":"D1FAE5","yellow":"D97706","yellow_lt":"FEF3C7",
    "red":"DC2626","red_lt":"FEE2E2",
}

def fill(h): return PatternFill("solid", fgColor=h)
def font(bold=False,size=10,color="0F172A",name="Arial"): return Font(bold=bold,size=size,color=color,name=name)
def border_thin():
    s=Side(style="thin",color=C["brd"]); return Border(top=s,bottom=s,left=s,right=s)
def border_med():
    s=Side(style="medium",color=C["navy"]); return Border(top=s,bottom=s,left=s,right=s)
def align(h="left",v="center",wrap=False,indent=0): return Alignment(horizontal=h,vertical=v,wrap_text=wrap,indent=indent)
def set_widths(ws,widths):
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
def hdr_row(ws,row,headers,bg=None):
    bg=bg or C["teal"]
    for i,h in enumerate(headers,1):
        c=ws.cell(row=row,column=i,value=h)
        c.fill=fill(bg);c.font=font(bold=True,size=9,color=C["white"]);c.alignment=align("center");c.border=border_thin()
    ws.row_dimensions[row].height=22
def dc(ws,row,col,value,alt=False,fmt=None,h="left",bold=False,fg=None,bg=None):
    c=ws.cell(row=row,column=col,value=value)
    c.fill=fill(bg or (C["row"] if alt else C["white"]))
    c.font=font(bold=bold,size=9,color=fg or C["dk"]);c.alignment=align(h);c.border=border_thin()
    if fmt: c.number_format=fmt
    return c
def sec(ws,row,text,cols,bg=None):
    ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}");c=ws[f"A{row}"]
    c.value=text;c.fill=fill(bg or C["hdr"]);c.font=font(bold=True,size=9,color=C["teal"])
    c.alignment=align("left",indent=1);ws.row_dimensions[row].height=16
def sem_col(label):
    l=(label or "").lower()
    if "cara" in l: return C["red"],C["red_lt"]
    if "revisar" in l: return C["yellow"],C["yellow_lt"]
    return C["green"],C["green_lt"]

def build_resumen(wb,datos):
    ws=wb.active;ws.title="Resumen Ejecutivo";ws.sheet_view.showGridLines=False
    ops=datos.get("operaciones",[]);ents=datos.get("entidades",[])
    deuda=sum(o.get("saldoPendiente",0) for o in ops if o.get("tipo")!="Tarjeta de Crédito")
    disp=sum(e.get("disponible",0) for e in ents)
    lt=sum(e.get("lineaTotal",0) for e in ents);uso=deuda/lt if lt else 0
    tn_n=sum(o.get("tna",0)*o.get("saldoPendiente",0) for o in ops if o.get("tipo")!="Tarjeta de Crédito")
    tn_d=sum(o.get("saldoPendiente",0) for o in ops if o.get("tipo")!="Tarjeta de Crédito")
    tna_p=tn_n/tn_d if tn_d else 0
    ops_act=len([o for o in ops if o.get("saldoPendiente",0)>0])
    ws.row_dimensions[1].height=8;ws.row_dimensions[2].height=40
    ws.merge_cells("A2:H2");c=ws["A2"]
    c.value="REPORTE FINANCIERO  ·  VALUE LATAM RADAR"
    c.fill=fill(C["navy"]);c.font=Font(bold=True,size=18,color=C["teal"],name="Arial")
    c.alignment=Alignment(horizontal="left",vertical="center",indent=2)
    ws.row_dimensions[3].height=5;ws.merge_cells("A3:H3");ws["A3"].fill=fill(C["teal"])
    for i,(lbl,val) in enumerate([("EMPRESA",datos.get("empresa","—")),("CUIT",datos.get("cuit","—")),
                                   ("FECHA DE CORTE",datos.get("fechaCorte","—")),("GENERADO POR","Value Latam Radar · valuelatam.com")],5):
        ws.row_dimensions[i].height=18
        ws.cell(row=i,column=1,value=lbl).font=font(bold=True,size=8,color=C["md"])
        ws.merge_cells(f"B{i}:D{i}");cv=ws[f"B{i}"];cv.value=val;cv.font=font(bold=True,size=10,color=C["dk"])
    ws.row_dimensions[10].height=8;ws.row_dimensions[11].height=5
    ws.merge_cells("A11:H11");ws["A11"].fill=fill(C["teal_lt"])
    kpis=[("DEUDA TOTAL",f"${int(deuda):,}".replace(",","."),C["navy"]),
          ("DISPONIBLE",f"${int(disp):,}".replace(",","."),C["teal"]),
          ("TNA PONDERADA",f"{tna_p:.1f}%",C["hdr"]),
          ("USO LÍNEAS",f"{uso*100:.1f}%",C["hdr"]),
          ("OPS. ACTIVAS",str(ops_act),C["hdr"])]
    ws.row_dimensions[13].height=20;ws.row_dimensions[14].height=30;ws.row_dimensions[15].height=5
    for idx,(lbl,val,bg) in enumerate(kpis,1):
        cl=ws.cell(row=13,column=idx,value=lbl);cl.fill=fill(bg);cl.font=font(bold=True,size=8,color=C["white"]);cl.alignment=align("center");cl.border=border_med()
        cv=ws.cell(row=14,column=idx,value=val);cv.fill=fill(bg);cv.font=Font(bold=True,size=15,color=C["white"],name="Arial");cv.alignment=align("center");cv.border=border_med()
        ws.cell(row=15,column=idx).fill=fill(C["teal"])
    ws.row_dimensions[17].height=5;ws.row_dimensions[18].height=20
    ws.merge_cells("A18:H18");c18=ws["A18"];c18.value="POSICIÓN POR ENTIDAD"
    c18.fill=fill(C["navy"]);c18.font=font(bold=True,size=11,color=C["teal"]);c18.alignment=align("left",indent=1)
    hdr_row(ws,19,["Entidad","Tipo","Calificada ($)","Tomada ($)","Disponible ($)","% Uso","Vto. Línea","Estado"])
    for i,e in enumerate(sorted(ents,key=lambda e:(e.get("tipoLinea","")!="Rotativa",-e.get("lineaTotal",0)))):
        r=20+i;ws.row_dimensions[r].height=18;alt=i%2==0
        lt_e=e.get("lineaTotal",0);ut_e=e.get("lineaUtilizada",0);d_e=e.get("disponible",0)
        p_e=ut_e/lt_e if lt_e else 0;venc=e.get("lineaVencida",False)
        est="VENCIDA" if venc else ("ALTO USO" if p_e>0.8 else ("MEDIO" if p_e>0.5 else "OK"))
        dc(ws,r,1,e.get("nombre",""),alt,bold=True)
        dc(ws,r,2,"↺ Rotativa" if e.get("tipoLinea")=="Rotativa" else f"◆ {e.get('subtipo','Puntual')}",alt,h="center")
        dc(ws,r,3,lt_e,alt,h="right",fmt='#,##0');dc(ws,r,4,ut_e,alt,h="right",fmt='#,##0')
        dc(ws,r,5,d_e,alt,h="right",fmt='#,##0',fg=C["red"] if d_e==0 else C["green"],bold=True)
        pc=dc(ws,r,6,p_e,alt,h="center",fmt='0.0%')
        if venc or p_e>0.8: pc.fill=fill(C["red_lt"]);pc.font=font(bold=True,size=9,color=C["red"])
        elif p_e>0.5: pc.fill=fill(C["yellow_lt"]);pc.font=font(bold=True,size=9,color=C["yellow"])
        else: pc.fill=fill(C["green_lt"]);pc.font=font(bold=True,size=9,color=C["green"])
        dc(ws,r,7,e.get("vtoLinea","—"),alt,h="center")
        sc=dc(ws,r,8,est,alt,h="center",bold=True)
        if est in["VENCIDA","ALTO USO"]: sc.fill=fill(C["red_lt"]);sc.font=font(bold=True,size=9,color=C["red"])
        elif est=="MEDIO": sc.fill=fill(C["yellow_lt"]);sc.font=font(bold=True,size=9,color=C["yellow"])
        else: sc.fill=fill(C["green_lt"]);sc.font=font(bold=True,size=9,color=C["green"])
    set_widths(ws,[30,16,16,16,16,10,14,12])

def build_operaciones(wb,datos):
    ws=wb.create_sheet("Operaciones");ws.sheet_view.showGridLines=False
    ops=datos.get("operaciones",[])
    ws.row_dimensions[1].height=35;ws.merge_cells("A1:Q1");c=ws["A1"]
    c.value="DETALLE DE OPERACIONES";c.fill=fill(C["navy"]);c.font=Font(bold=True,size=14,color=C["teal"],name="Arial")
    c.alignment=Alignment(horizontal="left",vertical="center",indent=2)
    ws.row_dimensions[2].height=4;ws.merge_cells("A2:Q2");ws["A2"].fill=fill(C["teal"])
    hdr_row(ws,3,["Entidad","Tipo","Nro. Ref.","Monto Orig. ($)","Saldo Pend. ($)","TNA %","Tipo Tasa","Plazo (m)","Desembolso","Cuotas Tot.","Cuotas Pend.","Próx. Vto.","Próx. Cuota ($)","Capital ($)","Interés ($)","Estado TNA","Diferencia (pp)"])
    rots=[o for o in ops if o.get("esRotativa") and o.get("tipo") not in ["Descubierto","Tarjeta de Crédito","Préstamo de Inversión","Leasing"]]
    punts=[o for o in ops if not o.get("esRotativa") or o.get("tipo") in ["Descubierto","Tarjeta de Crédito","Préstamo de Inversión","Leasing"]]
    r=4
    for grupo,ops_g in [("↺ OPERACIONES ROTATIVAS",rots),("◆ OPERACIONES PUNTUALES",punts)]:
        if not ops_g: continue
        sec(ws,r,grupo,17);r+=1
        for i,op in enumerate(ops_g):
            ws.row_dimensions[r].height=18;alt=i%2==0
            tna=op.get("tna",0) or 0;diff=op.get("diferencia",0) or 0
            sfg,sbg=sem_col(op.get("semaforo","OK"))
            dc(ws,r,1,op.get("entidad",""),alt,bold=True);dc(ws,r,2,op.get("tipo",""),alt)
            dc(ws,r,3,op.get("nroPrestamo","—"),alt,h="center")
            dc(ws,r,4,op.get("montoOrigen",0),alt,h="right",fmt='#,##0')
            dc(ws,r,5,op.get("saldoPendiente",0),alt,h="right",fmt='#,##0')
            dc(ws,r,6,tna/100,alt,h="center",fmt='0.00%')
            dc(ws,r,7,op.get("tipoTasa","Fija"),alt,h="center");dc(ws,r,8,op.get("plazo",0),alt,h="center")
            dc(ws,r,9,op.get("desembolso","—"),alt,h="center");dc(ws,r,10,op.get("cuotasTotal",0),alt,h="center")
            dc(ws,r,11,op.get("cuotasPendientes",0),alt,h="center");dc(ws,r,12,op.get("proxVto","—"),alt,h="center")
            dc(ws,r,13,op.get("proxCuotaTotal",0),alt,h="right",fmt='#,##0')
            dc(ws,r,14,op.get("proxCapital",0),alt,h="right",fmt='#,##0')
            dc(ws,r,15,op.get("proxInteres",0),alt,h="right",fmt='#,##0')
            dc(ws,r,16,op.get("semaforo","OK"),alt,h="center",bold=True,fg=sfg,bg=sbg)
            xc=dc(ws,r,17,diff,alt,h="center",fmt='+0.0;-0.0;-')
            if diff>5: xc.font=font(bold=True,size=9,color=C["red"])
            elif diff>2: xc.font=font(bold=True,size=9,color=C["yellow"])
            else: xc.font=font(bold=True,size=9,color=C["green"])
            r+=1
    ws.row_dimensions[r].height=20
    ws.cell(row=r,column=1,value="TOTALES").fill=fill(C["navy"]);ws.cell(row=r,column=1).font=font(bold=True,size=10,color=C["white"]);ws.cell(row=r,column=1).border=border_med()
    for col,field in [(4,"montoOrigen"),(5,"saldoPendiente"),(13,"proxCuotaTotal"),(14,"proxCapital"),(15,"proxInteres")]:
        tc=ws.cell(row=r,column=col,value=sum(o.get(field,0) or 0 for o in ops))
        tc.fill=fill(C["navy"]);tc.font=font(bold=True,size=10,color=C["white"]);tc.alignment=align("right");tc.number_format='#,##0';tc.border=border_med()
    set_widths(ws,[26,20,14,16,16,10,14,9,13,10,11,13,15,14,14,13,13]);ws.freeze_panes="A4"

def build_concentracion(wb,datos):
    ws=wb.create_sheet("Concentración Vencimientos");ws.sheet_view.showGridLines=False
    conc=datos.get("concentracion",[]);total_c=sum(c.get("total",0) for c in conc)
    ws.row_dimensions[1].height=35;ws.merge_cells("A1:G1");c=ws["A1"]
    c.value="CONCENTRACIÓN DE VENCIMIENTOS";c.fill=fill(C["navy"]);c.font=Font(bold=True,size=14,color=C["teal"],name="Arial")
    c.alignment=Alignment(horizontal="left",vertical="center",indent=2)
    ws.row_dimensions[2].height=4;ws.merge_cells("A2:G2");ws["A2"].fill=fill(C["teal"])
    sec(ws,4,"POR SEMANA (próximos 90 días)",5)
    hdr_row(ws,5,["Semana","Monto ($)","Cuotas","% del Total","Distribución"])
    bar_cols=[C["red"],C["yellow"],"2563EB",C["teal"]]
    for i,sem in enumerate(conc):
        r=6+i;ws.row_dimensions[r].height=20;alt=i%2==0
        monto=sem.get("total",0);cant=sem.get("cantidad",0);p=monto/total_c if total_c else 0
        dc(ws,r,1,sem.get("semana",f"Semana {i+1}"),alt);dc(ws,r,2,monto,alt,h="right",fmt='#,##0')
        dc(ws,r,3,cant,alt,h="center");dc(ws,r,4,p,alt,h="center",fmt='0.0%')
        bc=ws.cell(row=r,column=5,value="█"*int(p*25))
        bc.font=Font(color=bar_cols[i%len(bar_cols)],name="Arial",size=10)
        bc.fill=fill(C["row"] if alt else C["white"]);bc.border=border_thin()
    r_tot=6+len(conc)
    for col,val,fmt in [(1,"TOTAL",None),(2,total_c,'#,##0'),(3,sum(c.get("cantidad",0) for c in conc),None),(4,1.0,'0%')]:
        tc=ws.cell(row=r_tot,column=col,value=val);tc.fill=fill(C["navy"]);tc.font=font(bold=True,size=9,color=C["white"])
        tc.alignment=align("right" if col>1 else "left")
        if fmt: tc.number_format=fmt
        tc.border=border_med()
    chart=BarChart();chart.type="col";chart.title="Vencimientos por Semana"
    chart.y_axis.title="Monto ($)";chart.style=10;chart.width=18;chart.height=12;chart.legend=None
    chart.add_data(Reference(ws,min_col=2,min_row=5,max_row=5+len(conc)),titles_from_data=True)
    chart.set_categories(Reference(ws,min_col=1,min_row=6,max_row=5+len(conc)))
    ws.add_chart(chart,"G4")
    por_ent=datos.get("concentracionPorEntidad",[]);total_e=sum(e.get("total",0) for e in por_ent)
    es=r_tot+3;sec(ws,es,"POR ENTIDAD",4);hdr_row(ws,es+1,["Entidad","Monto ($)","% del Total",""])
    for i,e in enumerate(sorted(por_ent,key=lambda x:-x.get("total",0))):
        r=es+2+i;ws.row_dimensions[r].height=18;alt=i%2==0;monto=e.get("total",0);p=monto/total_e if total_e else 0
        dc(ws,r,1,e.get("nombre",""),alt,bold=True);dc(ws,r,2,monto,alt,h="right",fmt='#,##0');dc(ws,r,3,p,alt,h="center",fmt='0.0%')
    set_widths(ws,[30,18,12,30,2,2,2])

def build_entidades(wb,datos):
    ws=wb.create_sheet("Entidades");ws.sheet_view.showGridLines=False
    ents=datos.get("entidades",[])
    ws.row_dimensions[1].height=35;ws.merge_cells("A1:H1");c=ws["A1"]
    c.value="LÍNEAS DE CRÉDITO POR ENTIDAD";c.fill=fill(C["navy"]);c.font=Font(bold=True,size=14,color=C["teal"],name="Arial")
    c.alignment=Alignment(horizontal="left",vertical="center",indent=2)
    ws.row_dimensions[2].height=4;ws.merge_cells("A2:H2");ws["A2"].fill=fill(C["teal"])
    hdr_row(ws,3,["Entidad","Tipo Línea","Subtipo","Calificada ($)","Tomada ($)","Disponible ($)","% Uso","Vto. Línea"])
    r=4;last_t=None
    for i,e in enumerate(sorted(ents,key=lambda e:(e.get("tipoLinea","")!="Rotativa",-e.get("lineaTotal",0)))):
        t=e.get("tipoLinea","Puntual")
        if t!=last_t:
            sec(ws,r,"↺ LÍNEAS ROTATIVAS" if t=="Rotativa" else "◆ LÍNEAS PUNTUALES",8);r+=1;last_t=t
        ws.row_dimensions[r].height=18;alt=i%2==0
        lt=e.get("lineaTotal",0);ut=e.get("lineaUtilizada",0);d=e.get("disponible",0);p=ut/lt if lt else 0;venc=e.get("lineaVencida",False)
        dc(ws,r,1,e.get("nombre",""),alt,bold=True);dc(ws,r,2,t,alt,h="center");dc(ws,r,3,e.get("subtipo",""),alt,h="center")
        dc(ws,r,4,lt,alt,h="right",fmt='#,##0');dc(ws,r,5,ut,alt,h="right",fmt='#,##0')
        dc(ws,r,6,d,alt,h="right",fmt='#,##0',fg=C["red"] if d==0 else C["green"],bold=True)
        pc=dc(ws,r,7,p,alt,h="center",fmt='0.0%')
        if venc or p>0.8: pc.fill=fill(C["red_lt"]);pc.font=font(bold=True,size=9,color=C["red"])
        elif p>0.5: pc.fill=fill(C["yellow_lt"]);pc.font=font(bold=True,size=9,color=C["yellow"])
        else: pc.fill=fill(C["green_lt"]);pc.font=font(bold=True,size=9,color=C["green"])
        vc=dc(ws,r,8,e.get("vtoLinea","—"),alt,h="center")
        if venc: vc.fill=fill(C["red_lt"]);vc.font=font(bold=True,size=9,color=C["red"])
        r+=1
    ws.row_dimensions[r].height=20
    ws.cell(row=r,column=1,value="TOTALES").fill=fill(C["navy"]);ws.cell(row=r,column=1).font=font(bold=True,size=10,color=C["white"]);ws.cell(row=r,column=1).border=border_med()
    for col,field in [(4,"lineaTotal"),(5,"lineaUtilizada"),(6,"disponible")]:
        tc=ws.cell(row=r,column=col,value=sum(e.get(field,0) for e in ents))
        tc.fill=fill(C["navy"]);tc.font=font(bold=True,size=10,color=C["white"]);tc.alignment=align("right");tc.number_format='#,##0';tc.border=border_med()
    ws.freeze_panes="A4";set_widths(ws,[30,14,18,18,16,16,10,14])


@https_fn.on_request(cors=https_fn.options.CorsOptions(cors_origins="*",cors_methods=["POST","OPTIONS"]))
def generar_reporte_excel(req: https_fn.Request) -> https_fn.Response:
    try:
        datos=req.get_json(force=True)
        if not datos:
            return https_fn.Response(json.dumps({"error":"Sin datos"}),status=400,content_type="application/json")
        wb=Workbook()
        build_resumen(wb,datos);build_operaciones(wb,datos);build_concentracion(wb,datos);build_entidades(wb,datos)
        wb.properties.title=f"Reporte — {datos.get('empresa','')}";wb.properties.creator="Value Latam Radar"
        output=io.BytesIO();wb.save(output);output.seek(0)
        empresa=datos.get("empresa","Empresa").replace(" ","_").replace(".","")
        fecha=datos.get("fechaCorte","").replace("/","-")
        nombre=f"Reporte_{empresa}_{fecha}.xlsx"
        return https_fn.Response(output.read(),status=200,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition":f'attachment; filename="{nombre}"',"Access-Control-Allow-Origin":"*"})
    except Exception as e:
        return https_fn.Response(json.dumps({"error":str(e)}),status=500,content_type="application/json")
